from datetime import datetime, timedelta, date
from io import BytesIO

from django.db.models import Sum, Count, Q
from django.utils import timezone
from django.http import HttpResponse
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.pos.models import Order, OrderItem
from apps.products.models import Product


class DashboardStatsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if request.user.role != 'admin':
            return Response({'detail': 'Hanya admin.'}, status=403)

        today = timezone.now().date()
        today_start = timezone.make_aware(datetime.combine(today, datetime.min.time()))
        today_end = timezone.make_aware(datetime.combine(today, datetime.max.time()))

        paid_today = Order.objects.filter(
            status='paid',
            created_at__range=(today_start, today_end),
        )

        today_sales = paid_today.aggregate(total=Sum('grand_total'))['total'] or 0
        today_orders = paid_today.count()
        today_customers = paid_today.exclude(
            Q(customer_name='') | Q(customer_name__isnull=True)
        ).values('customer_name').distinct().count()

        total_products = Product.objects.filter(is_active=True).count()
        low_stock_count = Product.objects.filter(
            is_active=True, stock__lte=models.F('min_stock')
        ).count()

        recent_orders = Order.objects.select_related('cashier').prefetch_related(
            'items__product'
        ).order_by('-created_at')[:10]
        from .serializers import OrderSerializer
        recent_orders_data = OrderSerializer(recent_orders, many=True).data

        top_products = (
            OrderItem.objects.filter(order__status='paid')
            .values('product_id', 'product_name')
            .annotate(total_qty=Sum('quantity'))
            .order_by('-total_qty')[:10]
        )

        sales_by_date = []
        for i in range(6, -1, -1):
            d = today - timedelta(days=i)
            d_start = timezone.make_aware(datetime.combine(d, datetime.min.time()))
            d_end = timezone.make_aware(datetime.combine(d, datetime.max.time()))
            day_total = Order.objects.filter(
                status='paid', created_at__range=(d_start, d_end)
            ).aggregate(total=Sum('grand_total'))['total'] or 0
            sales_by_date.append({
                'date': d.isoformat(),
                'total': day_total,
            })

        return Response({
            'today_sales': today_sales,
            'today_orders': today_orders,
            'today_customers': today_customers,
            'total_products': total_products,
            'low_stock_count': low_stock_count,
            'recent_orders': recent_orders_data,
            'top_products': list(top_products),
            'sales_by_date': sales_by_date,
        })


import django.db.models as models
import django.db.models.functions as funcs


class SalesReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if request.user.role != 'admin':
            return Response({'detail': 'Hanya admin.'}, status=403)

        from_date = request.query_params.get('from_date', (timezone.now() - timedelta(days=30)).date().isoformat())
        to_date = request.query_params.get('to_date', timezone.now().date().isoformat())

        orders = Order.objects.filter(
            status='paid',
            created_at__date__gte=from_date,
            created_at__date__lte=to_date,
        ).order_by('-created_at')

        export = request.query_params.get('export')
        if export == 'excel':
            return self.export_excel(orders, from_date, to_date)
        if export == 'pdf':
            return self.export_pdf(orders, from_date, to_date)

        from .serializers import OrderSerializer
        serializer = OrderSerializer(orders, many=True)
        total_sales = orders.aggregate(total=Sum('grand_total'))['total'] or 0
        total_orders = orders.count()

        payment_breakdown = (
            orders.values('payment_method')
            .annotate(total=Sum('grand_total'), count=Count('id'))
            .order_by('-total')
        )

        return Response({
            'from_date': from_date,
            'to_date': to_date,
            'total_sales': total_sales,
            'total_orders': total_orders,
            'payment_breakdown': list(payment_breakdown),
            'orders': serializer.data,
        })

    def export_excel(self, orders, from_date, to_date):
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = 'Laporan Penjualan'

        ws['A1'] = 'LAPORAN PENJUALAN'
        ws.merge_cells('A1:H1')
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f'Periode: {from_date} s/d {to_date}'
        ws.merge_cells('A2:H2')

        headers = ['No', 'ID Order', 'Kasir', 'Pelanggan', 'Total', 'Metode', 'Tanggal']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True)

        for i, order in enumerate(orders, 1):
            row = i + 4
            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=order.id)
            ws.cell(row=row, column=3, value=order.cashier.username)
            ws.cell(row=row, column=4, value=order.customer_name or '-')
            ws.cell(row=row, column=5, value=int(order.grand_total))
            ws.cell(row=row, column=6, value=order.get_payment_method_display())
            ws.cell(row=row, column=7, value=order.created_at.strftime('%Y-%m-%d %H:%M'))

        total_row = len(orders) + 5
        ws.cell(row=total_row, column=4, value='TOTAL').font = Font(bold=True)
        total = orders.aggregate(total=Sum('grand_total'))['total'] or 0
        ws.cell(row=total_row, column=5, value=int(total)).font = Font(bold=True)

        for col in range(1, 8):
            ws.column_dimensions[chr(64 + col)].width = 20

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="laporan_penjualan_{from_date}_{to_date}.xlsx"'
        return response

    def export_pdf(self, orders, from_date, to_date):
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph(f'Laporan Penjualan', styles['Title']))
        elements.append(Paragraph(f'Periode: {from_date} s/d {to_date}', styles['Normal']))
        elements.append(Spacer(1, 20))

        data = [['No', 'Order ID', 'Kasir', 'Total', 'Metode', 'Tanggal']]
        for i, order in enumerate(orders, 1):
            data.append([
                i,
                str(order.id),
                order.cashier.username,
                f"Rp{int(order.grand_total):,}",
                order.get_payment_method_display(),
                order.created_at.strftime('%Y-%m-%d %H:%M'),
            ])

        if orders:
            total = orders.aggregate(total=Sum('grand_total'))['total'] or 0
            data.append(['', '', 'TOTAL', f'Rp{int(total):,}', '', ''])

        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))

        elements.append(table)
        doc.build(elements)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/pdf',
        )
        response['Content-Disposition'] = f'attachment; filename="laporan_penjualan_{from_date}_{to_date}.pdf"'
        return response
